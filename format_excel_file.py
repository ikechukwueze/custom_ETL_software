import os

from openpyxl import Workbook, load_workbook
		
from openpyxl.styles import Font, Color, Alignment, Border, Side, colors, numbers




def format_excel_file(excelfile, has_combined_stats=False, use_comma=False):

	#load excelfile
	workbook = load_workbook(excelfile)
	
	#get a list of all sheets in workbook
	worksheets = workbook.sheetnames

	for sheet in worksheets:
	#active_sheet = workbook.active
		active_sheet = workbook[sheet]

		if has_combined_stats:

			active_sheet.merge_cells('A1:A2')

			active_sheet['A1'] = 'Sites'

			active_sheet.delete_rows(3)


		# Create a few styles
		font = Font(name='Comic Sans MS', sz=10)
		center_aligned_text = Alignment(horizontal="center")
		thin_border = Side(border_style="thin")
		square_border = Border(top=thin_border, right=thin_border, bottom=thin_border, left=thin_border)

		comma_separator = numbers.BUILTIN_FORMATS[3]

		
		#add styles to a dictionary
		styles = {'font':font, 'alignment':center_aligned_text, 'border':square_border, 'comma_sep':comma_separator}

		
		



		#get the table dimension i.e occupied cell range e.g A1:D27
		table = active_sheet[active_sheet.dimensions]

		#iterate through each column then each row and apply styles
		for column in table:
			for row in column:
				
				row.font = styles['font']
				row.alignment = styles['alignment']
				row.border 	= styles['border']
				if use_comma:
					row.number_format = styles['comma_sep']


		            
	            
	workbook.save(filename=excelfile)
	os.startfile(excelfile)	